import os
import time
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import mimetypes


# Добавляем тип MIME для .webp
mimetypes.init()
mimetypes.add_type("image/webp", ".webp")


# Создаем временную папку для хранения изображений
temp_dir = 'temp_images'
os.makedirs(temp_dir, exist_ok=True)

# Инициализация веб-драйвера
driver = webdriver.Chrome()

# Функция для загрузки изображения
def download_image(url, filename):
    response = requests.get(url)
    if response.status_code == 200:
        with open(filename, 'wb') as f:
            f.write(response.content)
        return True
    else:
        return False



# Функция для записи ссылок в файл
def write_links_to_file(links):
    with open('links2.txt', 'a') as file:
        for link in links:
            file.write(link + '\n')

# Проверка наличия файла и загрузка существующих ссылок
existing_links = set()
if os.path.exists('links2.txt'):
    with open('links2.txt', 'r') as file:
        existing_links = set(file.read().splitlines())

links_to_process = list(existing_links)[:3]

# Создаем новую книгу Excel и активный лист
wb = Workbook()
ws = wb.active

# Заголовки столбцов
ws.append(["Ссылка на товар", "Наименование", "Цена в usd", "Brand", "Style", "Release Date", "Sole Material", "Upper Material", "Изображения"])

# Итерация по каждой ссылке
for link in links_to_process:
    driver.get(link)

    try:
        # Ожидание загрузки данных о товаре
        WebDriverWait(driver, 20).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME, 'GoodsDetail_wrapper__tZh3Z'))
        )
        time.sleep(5)

        # Получаем данные о товаре
        product_details = {}

        # Наименование товара
        product_name = driver.find_element(By.CLASS_NAME, 'MainInfo_title__YSsXk').text

        # Стоимость товара
        price_bar = driver.find_element(By.CLASS_NAME, 'MainInfo_priceBar__tcUgc')
        price_div = price_bar.find_element(By.TAG_NAME, 'div')
        product_price = price_div.text.replace('$', '')

        # Находим все элементы с классом 'ProductDetails_propertyItem__mGdzY'
        product_items = driver.find_elements(By.CLASS_NAME, 'ProductDetails_propertyItem__mGdzY')

        # Переменные для хранения значений
        product_brand = ""
        product_style = ""
        product_release_date = ""
        product_sole_material = ""
        product_upper_material = ""

        # Проходим по найденным элементам
        for item in product_items:
            # Находим элементы с классом 'ProductDetails_propertyLabel__ZlSsu'
            label = item.find_element(By.CLASS_NAME, 'ProductDetails_propertyLabel__ZlSsu').text
            # Находим элементы с классом 'ProductDetails_propertyValue__Aj_Cz'
            value = item.find_element(By.CLASS_NAME, 'ProductDetails_propertyValue__Aj_Cz').text
            # Проверяем, что текст элемента соответствует "Brand"
            if label == "Brand":
                product_brand = value
            # Проверяем, что текст элемента соответствует "Style"
            elif label == "Style":
                product_style = value
            elif label == "Release Date":
                product_release_date = str(value)
            elif label == "Sole Material":
                product_sole_material = str(value)
            elif label == "Upper Material":
                product_upper_material = str(value)


        # Получаем ссылки на изображения
        image_elements = driver.find_elements(By.XPATH, '//img[@class="PoizonImage_img__BNSaU ProductSkuImgs_img__gYd8t"]')  # Пример XPath

        # Список для хранения имен загруженных изображений
        image_names = []

        # Сохраняем изображения
        for index, image_element in enumerate(image_elements, start=1):
            image_url = image_element.get_attribute('src')
            # image_filename = f"{product_name}_{index}.jpg"
            image_filename = f"{product_name.replace(' ', '_')}_{index}.jpg"
            download_image(image_url, os.path.join(temp_dir, image_filename))
            print(f"Сохранено изображение: {image_filename}")  # Отладочный вывод
            image_names.append(image_filename)


        # Запись данных о товаре в Excel
        ws.append([link, product_name, product_price, product_brand, product_style, product_release_date, product_sole_material,
                   product_upper_material, ', '.join(image_names)])

    except TimeoutException as e:
        print(f"TimeoutException: {e}")
        print(f"Current page URL: {driver.current_url}")


# Добавляем изображения в книгу Excel
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=9, max_col=9):
    for cell in row:
        images = cell.value.split(', ')
        for image in images:
            image_path = os.path.join(temp_dir, image)
            if os.path.isfile(image_path):  # Проверяем, является ли путь файлом
                img = Image(image_path)
                img.width = img.width * 0.99
                img.height = img.height * 0.99
                img.anchor = cell.coordinate
                ws.add_image(img)
                cell.value = None


# Сохраняем книгу Excel
wb.save("product_details.xlsx")


# Удаляем временную папку с изображениями
for filename in os.listdir(temp_dir):
    file_path = os.path.join(temp_dir, filename)
    os.remove(file_path)
os.rmdir(temp_dir)


# Закрываем браузер
driver.quit()